"""The xPts sheet: assemble the inputs, score them in Python, write the workbook.

All nine of FPL's scoring elements are present.  The scoring itself lives in
xpts_calc.py; this file gathers the inputs, hands them over, and lays the
answer out in Excel.

    xPts season  =  rate x (xMins / 90)  +  appearance  +  DefCon  +  bonus
                   +  penalties

Only the first term is linear in minutes, which is why the season total is not
one multiplication.  Appearance and DefCon points are earned per match and
capped at the 38 matches a season has; penalties are a season count already,
because the taker's expected share prices how much of the season he plays.

Almost every cell is a value.  Six columns stay live, and they are all
arithmetic rather than modelling: the chain from `xMins 26/27` through matches,
appearance and DefCon points to the season total, VORP, and VORP per £m.  That
is the one knob worth turning during a draft -- change a player's minutes and
his value moves -- and it is cheap to keep honest.  Everything else changes by
re-running the model.

Blue-filled columns are inputs; grey ones come out of the model.
"""
from __future__ import annotations

import sys

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import xpts_calc
from common import OUT, PROC

INPUT_FILL = PatternFill("solid", fgColor="DDEBF7")     # editable
DERIVED_FILL = PatternFill("solid", fgColor="F2F2F2")   # from the data
HEAD_FILL = PatternFill("solid", fgColor="404040")
# Amber: a player whose club changed over the summer.  His clean sheets, saves
# and goals against follow him to the new club, but his xG and xA are still the
# rates he produced at the old one, and nothing in the data fixes that.
MOVED_FILL = PatternFill("solid", fgColor="FCE4D6")
MOVED_COLS = ("Player", "Club 25/26", "xG/90", "xA/90",
              "xG pts/90", "xA pts/90")

# Green: the club-by-club research moved this player's minutes off Solio's
# forecast.  Marked so that a disagreement is visible without reading two
# columns and subtracting.
RESEARCH_FILL = PatternFill("solid", fgColor="E2EFDA")
RESEARCH_COLS = ("xMins Solio adjusted", "Research delta",
                 "Research confidence", "Research reason")

# Scoring, straight from the FPL API's game_config for 2026/27.  Defined once,
# in xpts_calc; reproduced on the Assumptions sheet so the workbook says what it
# was scored under without anyone having to open the code.
SCORING = [(pos, *vals) for pos, vals in xpts_calc.SCORING.items()]

NOTES = [
    ("xMins", "Your minutes forecast for 2026/27, in the blue cell. Three "
              "columns sit behind it. 'xMins Solio' is Solio's raw forecast, "
              "which is forward-looking and so already prices injuries and "
              "late World Cup returns. 'xMins Solio adjusted' is that figure "
              "after the club-by-club research moved it, for the 226 players "
              "the research reached; everyone else keeps Solio's number, or "
              "2025/26 minutes where Solio has no view. 'xMins 26/27' starts "
              "at the adjusted figure and is yours to overwrite. 'xMins "
              "source' says which of the three a row is on, 'Research delta' "
              "how far the research moved it, and 'Research reason' why, with "
              "the article or podcast it came from. This is the single biggest "
              "lever in the model and the one number no dataset can settle."),
    ("Research", "Green cells mark the 226 players whose minutes the "
                 "club-by-club research changed. It read what beat writers, "
                 "local press and club podcasts were saying in the fortnight "
                 "to 16 August 2026, one club at a time, and every change "
                 "carries a source URL and a date. Confidence is the "
                 "researcher's own: high, medium or low. It is a dated "
                 "snapshot, not a live feed -- it does not know anything that "
                 "happened after that date, and it cannot see January."),
    ("Saves", "Goalkeepers only, 1 point per 3 saves settled every match. "
              "'Saves/match' is the club's expected saves, derived from the "
              "goals it is expected to concede and the quality of the shots it "
              "concedes them from. The sheet turns that into points with a "
              "Poisson, because two saves in a match are worth nothing: over "
              "2025/26 that flooring was worth 0.213 points per save, not "
              "0.333."),
    ("Goals conceded", "Goalkeepers and defenders, -1 per 2 conceded, settled "
                       "every match. Same Poisson treatment: ignoring the "
                       "per-match floor makes the penalty 57% too harsh."),
    ("Cards", "-1 a yellow, -3 a red. Booking rate is a real player trait -- "
              "year-on-year correlation +0.47 over 766 repeat player-seasons -- "
              "so yellows are the player's own three-season rate shrunk toward "
              "his position. Reds are not: at one per 200 nineties an "
              "individual red rate is noise, so everyone gets the position's."),
    ("Penalties", "Kept out of xG entirely and credited back here, because "
                  "penalty duty moves between seasons and Cole Palmer's 2025/26 "
                  "expected goals were 44% penalties. 'Pens/season' is the "
                  "club's expected penalties times this player's share of them, "
                  "which follows the listed taker order and how much of the "
                  "season he is expected to play. Season-level, not per 90."),
    ("P(CS)", "Probability the player's club keeps a clean sheet in a given match. "
              "Defaults to the value implied by season-long betting markets: "
              "bet365 and Spreadex points totals, converted to goal difference, "
              "then to goals for and against, then to expected goals in each of "
              "38 fixtures, then Poisson. Covers all 20 clubs including the "
              "promoted three. Validated on five seasons: the ordering across "
              "clubs is sound, the league-wide level carries about +/-13%."),
    ("Assist uplift", "FPL awards assists more loosely than Opta (rebounds, "
                      "deflections, won penalties). Across 2025/26 that was 765 "
                      "FPL assists vs 580 Opta assists, so 1.32. Both xG models "
                      "measure the Opta definition, so xA needs this to become "
                      "FPL assists. Applied in the model, not in the sheet: "
                      "changing the number here does nothing until scripts are "
                      "re-run."),
    ("xG source", "Adjusted xGOT where a player has shot-placement history, "
                  "otherwise blended xG. Column 'xG basis' says which was used. "
                  "Either way it is scaled to the non-penalty share of his "
                  "2025/26 expected goals, so penalties are counted once, in "
                  "the penalty column, and not twice."),
    ("DefCon", "Not a frozen rate. 'DefCon actions/90' is the player's underlying "
               "rate of qualifying defensive actions, shrunk toward the position "
               "mean for thin samples. The model turns that into 'DefCon hit %' "
               "with a negative binomial over 'Mins per start' against the "
               "threshold on the table above -- so if you change minutes per "
               "start, the hit rate moves with it. A defender on 10 actions/90 "
               "clears a threshold of 10 in 19% of 60-minute starts and 49% of "
               "90-minute starts. Both minutes inputs matter: xMins sets how "
               "many starts he gets, mins per start sets how likely each one "
               "pays. Negative binomial rather than Poisson because a player's "
               "actions are not a constant-rate process -- measured within "
               "player across starts, the variance is 1.5x the mean -- and "
               "because P(clearing the threshold) is convex in the rate, that "
               "spread produces MORE crossings, not fewer. A Poisson returned "
               "16% less DefCon than 2025/26 actually awarded and 12.4% less "
               "out of sample; the negative binomial returns -1.9% and -0.0%. "
               "The dispersion is fitted per position, two-fold, in "
               "defcon_model.py: 8 for defenders, 16 for midfielders."),
    ("Mins per start", "Minutes played in matches he started, over starts, "
                       "from 2025/26 -- with sendings-off dropped from both "
                       "sides, since a red card is not rotation risk. It is "
                       "shrunk toward the position mean by 10 starts, which "
                       "out of sample beats both the raw average (MAE 4.07 "
                       "against 5.20) and the position mean alone (4.76). "
                       "'Mins/start starts' is how many starts it rests on: a "
                       "small number means the figure is mostly the prior. "
                       "Editable, because the research knows things the "
                       "measurement cannot -- a player promoted to nailed is "
                       "usually hooked less often than he was, and no row from "
                       "last season can say so. Note the direction: at a fixed "
                       "xMins, a LOWER figure here means more matches and so "
                       "more appearance points. The early hook costs you "
                       "through xMins, not through this column."),
    ("Bonus", "Won per match, not per minute: 3/2/1 is contested once per "
              "fixture. 'Base BPS/90' is the player's BPS with every event the "
              "model projects elsewhere — goals, assists, clean sheets, saves, "
              "goals conceded, cards — stripped out, leaving appearance, "
              "defensive actions, passing and carrying. That remainder is what "
              "repeats: across 2025/26's two halves, total BPS per 90 "
              "replicated at r=0.39 and base BPS per 90 at r=0.78, so the old "
              "column carried mostly noise and double-counted the goals. "
              "'Bonus/match' simulates one match — his base BPS, plus goals and "
              "assists drawn from his own team's goals, plus the clean sheet — "
              "against the three scores he actually had to beat, sampled from "
              "380 real fixtures and conditioned on the scoreline. That "
              "conditioning is why a Manchester City assist is worth less "
              "bonus than the same assist elsewhere: in a 4-0 win four "
              "team-mates and a clean sheet are competing for the same three "
              "points. Scaled to the bonus that actually exists, 6 points a "
              "fixture, by a factor of 1.06 to 1.34 depending on position."),
    ("VORP", "xPts season minus the replacement level at that position — the "
             "best player still on the board once every team has filled its "
             "slots. Set the league size and slots below. Raw xPts asks who "
             "scores most; VORP asks who scores most above what you would get "
             "at that slot anyway."),
    ("VORP per £m", "VORP divided by price. Pure VORP measures scarcity alone; "
                    "this measures scarcity against the £100m budget that "
                    "still binds you in a normal FPL season."),
    ("Appearance", "1 point for playing, 2 for 60 minutes or more, earned per "
                   "match rather than per 90 -- which is why it is shown as a "
                   "season total. 'Matches' is xMins divided by 'Mins per "
                   "start', capped at the 38 a season has. The cap matters: "
                   "Christian Nørgaard's 45 minutes per start came from a "
                   "cameo-heavy year, and 2,804 forecast minutes divided by it "
                   "claims 62 matches. Once capped he is making 38 longer "
                   "appearances instead, which clears the 60-minute line. 'App "
                   "pts 25/26' is what he actually earned last season, for "
                   "comparison only -- it does not feed the model."),
    ("Moved club", "Amber cells mark a player who changed club over the "
                   "summer, and mark exactly the numbers the move breaks. His "
                   "clean sheets, saves and goals against all follow him to the "
                   "new club correctly -- those are keyed on the club, never on "
                   "the player. His xG and xA do not: they are the rates he "
                   "produced at the old club, in the old side's shape, and "
                   "nothing in any dataset repairs that. Treat the amber "
                   "figures as needing your own judgement."),
    ("Live cells", "Almost everything on this sheet is a value, computed in "
                   "Python. Six columns are still real formulas -- Matches, App "
                   "pts season, DC pts season, xPts season, VORP and VORP per "
                   "£m -- so that changing 'xMins 26/27' or 'Mins per start' "
                   "moves a player's total and his rank the way it should. "
                   "Change anything else and you are editing a number the model "
                   "will overwrite on the next run."),
]


def _xg_basis(m: pd.DataFrame):
    """Which xG figure each player is scored on, and the pieces behind it.

    Split out of build_rows so that bonus_model can score its simulation on
    the same expected goals the rest of the model uses.  It used to run on the
    raw FPL figure, which counts penalties, so Erling Haaland's simulated bonus
    was priced off an xG a quarter of which is credited separately in the
    penalty column -- and then dragged back by a linear slope over a distance
    far too big for a first-order correction to be honest about.
    """
    ratio = m.get("placement_ratio")
    adj90 = m.get("adjusted_xGOT_p90")
    blend90 = m["xG_blend_p90"]
    use_adj = adj90.notna() & ratio.notna() & (m.get("hist_xG", 0).fillna(0) > 0)
    return ratio, adj90, blend90, use_adj


def model_rates(m: pd.DataFrame) -> pd.DataFrame:
    """The three inputs bonus depends on: non-penalty xG/90, xA/90 and P(CS).

    Exactly what build_rows scores with, so the bonus simulation and the rest
    of the model see the same player.
    """
    _, adj90, blend90, use_adj = _xg_basis(m)
    np_share = m.get("np_share")
    np_share = pd.Series(1.0, index=m.index) if np_share is None else np_share.fillna(1.0)
    out = pd.DataFrame(index=m.index)
    # Rounded, not clipped, and rounded to the same 4 places build_rows uses.
    # A clip here would silently put the bonus simulation on a different xG
    # from the one the player is scored on, and the slope would then have to
    # extrapolate across the gap -- which for a 200-minute sample with a
    # nonsense xG/90 means extrapolating a long way.
    out["xG_p90"] = (adj90.where(use_adj, blend90) * np_share).round(4)
    out["xA_p90"] = m["xA_blend_p90"].round(4)
    odds = PROC / "cs_from_odds.csv"
    if odds.exists():
        o = pd.read_csv(odds)
        out["pcs"] = m["team_2627"].map(dict(zip(o["team_short"], o["p_cs"])))
    else:
        out["pcs"] = np.nan
    return out


def build_rows() -> pd.DataFrame:
    m = pd.read_csv(PROC / "master_2025_26.csv")

    # Default P(CS): derived from season-long betting markets where available,
    # since last season's clean-sheet count was the weakest of the three
    # predictors tested (R2 0.14, against 0.22 for xGA).  The market prices this
    # summer's transfers and covers the promoted clubs, which no backward-looking
    # statistic can.  Falls back to last season's rate if the market file is
    # missing.  Keyed on the club, never on a player, so a defender who moved
    # does not import his old defence.
    odds_path = PROC / "cs_from_odds.csv"
    if odds_path.exists():
        o = pd.read_csv(odds_path)
        cs = dict(zip(o["team_short"], o["p_cs"].round(3)))
        print(f"  P(CS) default: market-implied for {len(cs)} clubs")
    else:
        teams = pd.read_csv(PROC / "understat_teams.csv")
        teams = teams[teams.season == "2025/26"]
        cs = dict(zip(teams["team_short"], teams["cs_rate"].round(3)))
        print("  P(CS) default: last season's clean-sheet rate (no market file)")

    d = pd.DataFrame({
        "player": m["player"],
        "team": m["team_2627"],
        "club_2526": m.get("club_2526"),
        "moved_club": m.get("moved_club", pd.Series(False, index=m.index)).fillna(False),
        "pos": m["position"],
        "pos_2526": m.get("pos_2526"),
        "price": m["price_2627"],
        "draftable": m["draftable_2627"],
        "status": m["status"],
        "news": m["news"],
        "mins_2526": m["minutes"],
        "pts_2526": m["fpl_total_points"],
    })

    # --- xG basis: adjusted xGOT where there is placement history ----------
    ratio, adj90, blend90, use_adj = _xg_basis(m)
    # Both models measure expected goals including penalties.  Penalties are
    # credited separately and explicitly, so they come out here first -- by the
    # player's own measured non-penalty share, from Understat's npxG.  The same
    # share is applied to the adjusted-xGOT figure, which assumes shot placement
    # is as good from open play as from the spot; that is close enough to true
    # and errs by a fraction of a goal.
    np_share = m.get("np_share")
    np_share = pd.Series(1.0, index=m.index) if np_share is None else np_share.fillna(1.0)
    d["xG_p90"] = (adj90.where(use_adj, blend90) * np_share).round(4)
    d["xG_basis"] = use_adj.map({True: "adjusted xGOT", False: "blended xG"})
    d.loc[d["xG_p90"].isna(), "xG_basis"] = ""
    d["np_share"] = np_share.round(3)
    d["placement_ratio"] = ratio.round(3)
    d["hist_xG"] = m.get("hist_xG").round(1)

    d["xA_p90"] = m["xA_blend_p90"].round(4)

    # --- DefCon: an action RATE plus minutes per start, not a frozen rate ---
    # The sheet turns these into P(clearing the threshold) with a live Poisson,
    # so the answer responds to minutes instead of ignoring them.
    d["defcon_lambda"] = m.get("defcon_lambda")
    d.loc[m["position"] == "GKP", "defcon_lambda"] = 0.0
    # Dispersion of the per-match action count, fitted per position in
    # defcon_model.py.  Keyed on the 2026/27 position, not the one carried in
    # from last season: the DefCon *threshold* follows the position a player is
    # scored at, so the distribution it is compared against has to as well.
    # Three converted midfielders were being scored as defenders against a
    # midfielder's dispersion, and the 58 summer arrivals with no DefCon row at
    # all were falling back to the Poisson the change exists to replace.
    sz = m.get("defcon_size")
    if sz is None:
        d["defcon_size"] = np.inf
    else:
        by_pos = (pd.DataFrame({"pos": m.get("pos_2526"), "sz": sz}).dropna()
                  .groupby("pos")["sz"].median())
        d["defcon_size"] = m["position"].map(by_pos).fillna(np.inf)
    # Anyone with no Premier League record -- a promoted club's keeper, an
    # incoming signing -- has no measured appearance length, so he gets the
    # default.  Leaving it blank makes his match count zero, and with it his
    # appearance and DefCon points, however many minutes he is forecast.
    mps = m.get("mins_per_start")
    if mps is None:
        # optional() drops a whole file's columns when it is missing, so the
        # symptom is a None where a Series is expected.  Name the step that
        # produces it rather than letting pandas raise on `.fillna`.
        raise SystemExit(
            "master has no `mins_per_start` -- run defcon_model.py before "
            "build_master.py (run_all.py does this in order).")
    d["mins_per_start"] = mps.fillna(xpts_calc.DEFAULT_MINS_PER_START)
    # How many starts the figure rests on, after sendings-off are removed.  It
    # is already shrunk toward the position mean by 10 starts, so a small count
    # here means the number is mostly the position prior, not this player.
    d["mps_starts"] = m.get("mps_starts")

    # --- bonus: won per match, from projected BPS, not carried forward -----
    # What used to sit here was last season's realised bonus per 90.  That
    # credited every player for the goals, assists and clean sheets he actually
    # got -- the three things this model already projects from scratch, and the
    # three that repeat least (total BPS/90 replicates half-to-half at r=0.39;
    # strip the events out and the remainder replicates at r=0.78).
    bpm = m.get("bonus_per_match")
    if bpm is None:
        raise SystemExit(
            "master has no `bonus_per_match` -- run bonus_model.py before "
            "build_master.py (run_all.py does this in order).")
    d["bonus_per_match"] = bpm
    d["bonus_d_xg"] = m.get("d_bonus_d_xg90")
    d["bonus_d_xa"] = m.get("d_bonus_d_xa90")
    d["bonus_d_pcs"] = m.get("d_bonus_d_pcs")
    d["bonus_xg_base"] = m.get("bonus_xg_base")
    d["bonus_xa_base"] = m.get("bonus_xa_base")
    d["bonus_pcs_base"] = m.get("bonus_pcs_base")
    d["base_bps_p90"] = m.get("base_bps_p90")

    # --- club defensive volume, and the three remaining scoring elements ----
    d["saves_per_match"] = m.get("saves_per_match")
    d["ga_per_match"] = m.get("ga_per_match")
    d["card_pts_p90"] = m.get("card_pts_p90")
    # Penalties are a season total, not a rate: the expected count already
    # prices how much of the season the taker is available for.
    d["pens_season"] = m.get("pens_expected")
    d["pen_save_pts"] = m.get("pen_save_pts")

    # --- appearance points, from minutes per appearance --------------------
    # Not a frozen per-90 rate.  Appearance points are earned per match, so a
    # substitute who plays 15 minutes at a time banks far more of them per 90
    # than a starter -- Christian Nørgaard's 101 minutes across seven cameos
    # worked out at 6.2 appearance points per 90.  Freezing that and scaling it
    # by a full season's projected minutes claims 187 appearances.  The sheet
    # instead computes appearances as xMins / mins per start, which is the same
    # input the DefCon model already uses, so the answer moves with the minutes
    # assumption instead of ignoring it.
    played, full60 = m.get("gw_played"), m.get("gw_full_60")
    d["app_pts_2526"] = (2 * full60 + (played - full60)).astype("Float64")

    # Solio's own minutes forecast, as an alternative to last season's actual.
    # It is the better starting point for anyone who was injured or came back
    # late from the World Cup, since it is forward-looking.
    d["solio_season_xmins"] = m.get("solio_season_xmins")
    d["xmins_pattern"] = m.get("xmins_pattern")
    # Solio's forecast is the default because it is forward-looking: it prices
    # this summer's injuries and the late World Cup returns, which last
    # season's minutes cannot.  Players Solio has no view on keep their 2025/26
    # minutes, and 'xMins source' records which of the three a row is using.
    solio = m.get("solio_season_xmins")
    base = solio.where(solio.notna(), m["minutes"])
    base_source = np.where(solio.notna(), "Solio", "25/26 actual")

    # The club-by-club research sits in its own column rather than overwriting
    # Solio's.  Both are visible, so the size of every disagreement can be read
    # off the sheet -- and if the research turns out to be wrong about a player,
    # the number it replaced is still right there.
    research = m.get("research_xmins")
    if research is None:
        research = pd.Series(pd.NA, index=m.index, dtype="Float64")
    d["research_xmins"] = research
    d["research_confidence"] = m.get("research_confidence")
    d["research_reason"] = m.get("research_reason")
    d["xmins_adjusted"] = research.where(research.notna(), base).round(0)
    d["research_delta"] = (d["xmins_adjusted"] - base).round(0)

    # The editable cell defaults to the researched figure: it is the more
    # informed of the two wherever it exists.  Overwrite the blue cell to
    # disagree -- that is what it is for.
    d["xMins_input"] = d["xmins_adjusted"]
    d["xmins_source"] = np.where(research.notna(), "research", base_source)
    d.loc[d["xMins_input"].isna(), "xmins_source"] = ""
    d["pcs_input"] = m["team_2627"].map(cs)

    d = d[d["draftable"] | d["mins_2526"].notna()]
    return d.sort_values("pts_2526", ascending=False, na_position="last")


COLUMNS = [
    # (header, source column or None, kind)
    ("Player", "player", "text"),
    ("Team 26/27", "team", "text"),
    ("Club 25/26", "club_2526", "text"),
    ("Moved club", "moved_club", "text"),
    ("Pos", "pos", "text"),
    ("Pos 25/26", "pos_2526", "text"),
    ("Draftable", "draftable", "text"),
    ("Price", "price", "num"),
    ("Status", "status", "text"),
    ("News", "news", "text"),
    ("Mins 25/26", "mins_2526", "num"),
    ("Pts 25/26", "pts_2526", "num"),
    ("xMins Solio", "solio_season_xmins", "derived"),
    ("xMins Solio adjusted", "xmins_adjusted", "derived"),
    ("Research delta", "research_delta", "derived"),
    ("Research confidence", "research_confidence", "text"),
    ("xMins pattern", "xmins_pattern", "text"),
    ("xMins 26/27", "xMins_input", "input"),
    ("xMins source", "xmins_source", "text"),
    ("Research reason", "research_reason", "text"),
    # Sits with the minutes block rather than with DefCon: it is how long a
    # start lasts, which is the first thing you want next to a minutes forecast.
    # Editable, because the research knows things the measurement cannot: a
    # player promoted to nailed is usually hooked less often than he was, and
    # nothing in last season's rows can say so.
    ("Mins per start", "mins_per_start", "input"),
    ("Mins/start starts", "mps_starts", "num"),
    ("Matches", None, "formula"),
    ("P(CS)", "pcs_input", "input"),
    ("xG/90", "xG_p90", "derived"),
    ("xG basis", "xG_basis", "text"),
    ("Non-pen share", "np_share", "derived"),
    ("Placement ratio", "placement_ratio", "derived"),
    ("Placement sample xG", "hist_xG", "derived"),
    ("xA/90", "xA_p90", "derived"),
    ("DefCon actions/90", "defcon_lambda", "derived"),
    ("DefCon hit %", "defcon_hit", "derived"),
    ("Base BPS/90", "base_bps_p90", "derived"),
    # The priced figure, not the raw one out of bonus_model: it is the raw
    # figure moved onto this sheet's P(CS), which is market-implied and so a
    # little above the Poisson the simulation drew clean sheets from.  Writing
    # the raw one would leave the sheet and the model disagreeing by up to
    # 0.7 points a season for no reason a reader could see.
    ("Bonus/match", "bonus_per_match_priced", "derived"),
    ("App pts 25/26", "app_pts_2526", "derived"),
    ("Saves/match", "saves_per_match", "derived"),
    ("Goals against/match", "ga_per_match", "derived"),
    ("Card rate/90", "card_pts_p90", "derived"),
    ("Pens/season", "pens_season", "input"),
    ("Pen save pts", "pen_save_pts", "derived"),
    # --- the seven elements that scale with minutes, and their sum ----------
    ("xG pts/90", "xg_pts_p90", "derived"),
    ("xA pts/90", "xa_pts_p90", "derived"),
    ("CS pts/90", None, "formula"),
    ("Save pts/90", "save_pts_p90", "derived"),
    ("GC pts/90", "gc_pts_p90", "derived"),
    ("Cards pts/90", "cards_pts_p90", "derived"),
    ("Rate pts/90", None, "formula"),
    # --- the two earned per match, plus penalties, as season totals ---------
    ("App pts season", None, "formula"),
    ("DC pts season", None, "formula"),
    ("Bonus pts season", None, "formula"),
    ("Pen pts season", None, "formula"),
    ("xPts season", None, "formula"),
    ("xPts/90", None, "formula"),
    ("VORP", None, "formula"),
    ("VORP per £m", None, "formula"),
    # Helper columns: xPts season, but blank unless the player is draftable and
    # plays that position. LARGE() over one of these gives the replacement
    # level without needing an array formula, which keeps the sheet portable.
    ("_pool GKP", None, "helper"),
    ("_pool DEF", None, "helper"),
    ("_pool MID", None, "helper"),
    ("_pool FWD", None, "helper"),
    # This row's scoring multipliers, written as values so the live formulas
    # above need no VLOOKUP into the Assumptions sheet.
    ("_app 60+", "app_mult", "helper"),
    ("_dc pts", "dc_mult", "helper"),
    ("_cs pts", "cs_mult", "helper"),
    ("_goal pts", "goal_mult", "helper"),
]

POSITIONS = ["GKP", "DEF", "MID", "FWD"]
SQUAD_SLOTS = {"GKP": 2, "DEF": 5, "MID": 5, "FWD": 3}   # an FPL 15-man squad
LEAGUE_TEAMS = 8
# Measured over 2022/23-2025/26: 381 penalties taken, 316 scored.
PENALTY_CONVERSION = 0.8294
# First row of the four-position VORP table on the Assumptions sheet.  It sits
# below the notes, so it moves whenever a note is added or reworded.
#
# Derived rather than hardcoded.  The constant it replaces was 36, which with
# 15 notes already put the notes loop on top of the "VORP settings" heading and
# silently erased it -- the notes are written after the VORP block, so they win.
# Two more notes pushed them onto "Teams in league" as well, which is what
# verify_xpts caught.  The +5 is the blank row plus the four-row settings
# header, so the table always clears the last note.
NOTES_ROW = 18                  # first note; "Notes" heading sits above it
VORP_ROW = NOTES_ROW + len(NOTES) + 5
TEAMS_ROW = VORP_ROW - 2        # the "Teams in league" input, referenced by LARGE()
# Season totals read better to one decimal than to three.
SEASON_FORMAT = {"Matches": "0.0", "xMins Solio adjusted": "0",
                 "Research delta": "0", "App pts season": "0.0",
                 "DC pts season": "0.0", "Pen pts season": "0.0",
                 "xPts season": "0.0", "VORP": "0.0", "Mins per start": "0.0"}


def main() -> int:
    from openpyxl import load_workbook

    d = build_rows()

    # Score in Python, then write the answers out.  The sheet keeps the live
    # chain from xMins onward, so the columns it recomputes are written as
    # formulas below and the values here are only used for the summary print.
    scored = xpts_calc.score(d)
    d = d.join(scored)
    mult = xpts_calc.multipliers(d["pos"])
    d["app_mult"] = mult["appearance"]
    d["dc_mult"] = mult["defcon"]
    d["cs_mult"] = mult["clean_sheet"]
    d["goal_mult"] = mult["goal"]

    path = OUT / "FPL_2026_27_draft_data.xlsx"
    if not path.exists():
        print(f"!! {path} not found -- run export_excel.py first", file=sys.stderr)
        return 1
    wb = load_workbook(path)

    for name in ("xPts model", "Assumptions"):
        if name in wb.sheetnames:
            del wb[name]

    # ---------------- Assumptions ----------------
    a = wb.create_sheet("Assumptions", 1)
    a["A1"] = "Scoring (FPL 2026/27, read from the game's own config)"
    a["A1"].font = Font(bold=True)
    a.append([])
    a.append(["Pos", "Goal", "Assist", "Clean sheet", "DefCon", "Appearance 60+",
              "DefCon threshold", "Saves per point", "Conceded per -1"])
    for row in SCORING:
        a.append(list(row))
    for c in a[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL

    a["A10"] = "Global inputs"
    a["A10"].font = Font(bold=True)
    a["A11"] = "Assist uplift (FPL assists / Opta assists)"
    a["B11"] = xpts_calc.ASSIST_UPLIFT
    a["A12"] = "DefCon points per qualifying match"
    a["B12"] = 2
    a["A13"] = "Penalty conversion rate"
    a["B13"] = PENALTY_CONVERSION
    a["B13"].fill = INPUT_FILL
    a["A14"] = "Points for a missed penalty"
    a["B14"] = xpts_calc.MISS_POINTS
    a["A15"] = "Matches in a season"
    a["B15"] = xpts_calc.MATCHES

    # ---- VORP settings, below the notes so nothing above shifts ----------
    head = VORP_ROW - 1
    a.cell(row=TEAMS_ROW - 2, column=1, value="VORP settings").font = Font(bold=True)
    a.cell(row=TEAMS_ROW, column=1, value="Teams in league")
    a.cell(row=TEAMS_ROW, column=2, value=LEAGUE_TEAMS).fill = INPUT_FILL
    for c, label in enumerate(("Position", "Squad slots per team",
                               "Replacement xPts"), start=1):
        cell = a.cell(row=head, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
    for i, pos in enumerate(POSITIONS):
        a.cell(row=VORP_ROW + i, column=1, value=pos)
        a.cell(row=VORP_ROW + i, column=2,
               value=SQUAD_SLOTS[pos]).fill = INPUT_FILL
        # column C is filled in below, once the helper columns have letters
    tail = VORP_ROW + len(POSITIONS) + 1
    a.cell(row=tail, column=1,
           value=("Replacement level is the best player at that position still "
                  "available once every team has filled its slots: the "
                  "(teams x slots + 1)-th best. Only players registered for "
                  "2026/27 count toward it.")).alignment = Alignment(
        wrap_text=True, vertical="top")
    a.row_dimensions[tail].height = 46

    a.cell(row=NOTES_ROW - 1, column=1, value="Notes").font = Font(bold=True)
    r = NOTES_ROW
    for label, text in NOTES:
        a.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = a.cell(row=r, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        a.row_dimensions[r].height = 46
        r += 1
    a.column_dimensions["A"].width = 16
    a.column_dimensions["B"].width = 105

    # ---------------- xPts model ----------------
    ws = wb.create_sheet("xPts model", 1)
    headers = [h for h, _, _ in COLUMNS]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="bottom")

    col = {h: get_column_letter(i + 1) for i, (h, _, _) in enumerate(COLUMNS)}

    for i, rec in enumerate(d.to_dict("records"), start=2):
        vals = []
        for h, src, kind in COLUMNS:
            if kind == "formula":
                vals.append(None)
            else:
                v = rec.get(src)
                vals.append(None if pd.isna(v) else v)
        ws.append(vals)

        xm = f"{col['xMins 26/27']}{i}"
        mps = f"{col['Mins per start']}{i}"

        # Clean sheets and the rate sum stay live so that editing P(CS) -- the
        # one modelling assumption a human might genuinely want to override for
        # a particular club -- still moves the total.
        ws[f"{col['CS pts/90']}{i}"] = (
            f"=N({col['P(CS)']}{i})*N({col['_cs pts']}{i})")
        ws[f"{col['Rate pts/90']}{i}"] = (
            f"=SUM({col['xG pts/90']}{i}:{col['Cards pts/90']}{i})")

        # Matches, capped at the 38 a season has.  Without the cap a forecast
        # of 2,804 minutes against 45 minutes per start claims 62 of them.
        ws[f"{col['Matches']}{i}"] = (
            f"=IF(N({mps})<=0,0,MIN(Assumptions!$B$15,N({xm})/N({mps})))")
        mt = f"{col['Matches']}{i}"
        # 1 point an appearance, 2 if it lasts 60 minutes.  The 60-minute test
        # uses minutes per *appearance* recomputed from the capped count, not
        # the raw minutes per start: a player pinned at 38 matches is making
        # longer appearances than his history suggests, not more of them.
        # The "1 point" branch has to be gated on the multiplier too, or a player
        # with no 2026/27 position -- who scores nothing anywhere else -- would
        # still bank a point for every short appearance.
        ws[f"{col['App pts season']}{i}"] = (
            f"=IF(OR(N({mt})<=0,N({col['_app 60+']}{i})=0),0,"
            f"N({mt})*IF(N({xm})/N({mt})>=60,N({col['_app 60+']}{i}),1))")
        # 2 points every match the threshold is cleared.  The probability of
        # clearing it comes from the model; only the count of chances is live.
        ws[f"{col['DC pts season']}{i}"] = (
            f"=N({mt})*N({col['DefCon hit %']}{i})*N({col['_dc pts']}{i})")
        # Bonus is contested once per fixture, so it is won per match like
        # appearance and DefCon points -- not per minute.  'Bonus/match' is one
        # simulated match at this player's minutes per start; only the count of
        # matches is live here, which is the same treatment the other two get.
        # Gated on the appearance multiplier for the same reason App pts season
        # is: a player with no 2026/27 position has left the league and must
        # score nothing.  Bonus is the one element with no positional
        # multiplier of its own to zero him out.
        ws[f"{col['Bonus pts season']}{i}"] = (
            f"=IF(N({col['_app 60+']}{i})=0,0,N({mt})*N({col['Bonus/match']}{i}))")
        # Penalties are already an expected count for the whole season -- the
        # taker's share prices his availability -- so they are never scaled by
        # minutes.
        ws[f"{col['Pen pts season']}{i}"] = (
            f"=N({col['Pens/season']}{i})*(Assumptions!$B$13*N({col['_goal pts']}{i})"
            f"-(1-Assumptions!$B$13)*Assumptions!$B$14)"
            f"+N({col['Pen save pts']}{i})")
        ws[f"{col['xPts season']}{i}"] = (
            f"=N({col['Rate pts/90']}{i})*N({xm})/90"
            f"+N({col['App pts season']}{i})+N({col['DC pts season']}{i})"
            f"+N({col['Bonus pts season']}{i})+N({col['Pen pts season']}{i})")
        # Shown per 90 for comparability. Penalties are excluded: they do not
        # scale with minutes, so folding them in would flatter a part-time taker.
        ws[f"{col['xPts/90']}{i}"] = (
            f"=IF(N({xm})<=0,0,(N({col['xPts season']}{i})"
            f"-N({col['Pen pts season']}{i}))*90/N({xm}))")

        season = f"{col['xPts season']}{i}"
        for pos in POSITIONS:
            ws[f"{col['_pool ' + pos]}{i}"] = (
                f'=IF(AND(${col["Draftable"]}{i}=TRUE,${col["Pos"]}{i}="{pos}"),'
                f'{season},"")')
        ws[f"{col['VORP']}{i}"] = (
            f'=IFERROR({season}-VLOOKUP(${col["Pos"]}{i},'
            f'Assumptions!$A${VORP_ROW}:$C${VORP_ROW + 3},3,FALSE),"")')
        ws[f"{col['VORP per £m']}{i}"] = (
            f'=IFERROR({col["VORP"]}{i}/{col["Price"]}{i},"")')

        moved = bool(rec.get("moved_club"))
        researched = pd.notna(rec.get("research_xmins"))
        for h, _, kind in COLUMNS:
            cell = ws[f"{col[h]}{i}"]
            if moved and h in MOVED_COLS:
                cell.fill = MOVED_FILL
            elif researched and h in RESEARCH_COLS:
                cell.fill = RESEARCH_FILL
            elif kind == "input":
                cell.fill = INPUT_FILL
            elif kind in ("derived", "formula"):
                cell.fill = DERIVED_FILL
            if kind in ("num", "derived", "formula") and h != "Pts 25/26":
                cell.number_format = (
                    "0" if kind == "num" else SEASON_FORMAT.get(h, "0.000"))

    last = ws.max_row
    for i, pos in enumerate(POSITIONS):
        pool = f"'xPts model'!${col['_pool ' + pos]}$2:${col['_pool ' + pos]}${last}"
        r = VORP_ROW + i
        a.cell(row=r, column=3,
               value=f"=IFERROR(LARGE({pool},$B${TEAMS_ROW}*B{r}+1),0)"
               ).number_format = "0.0"
    for h, _, kind in COLUMNS:
        if kind == "helper":
            ws.column_dimensions[col[h]].hidden = True

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    widths = {"Player": 22, "News": 30, "xG basis": 14,
              "Placement sample xG": 12, "Appearance/90": 12}
    for h, _, _ in COLUMNS:
        ws.column_dimensions[col[h]].width = widths.get(h, 11)

    wb.save(path)
    print(f"added 'xPts model' ({ws.max_row - 1} players) and 'Assumptions' to {path}")
    print("  blue cells are yours to edit: xMins 26/27, Mins per start, "
          "P(CS), Pens/season")
    print("  everything else is a value from the model -- change it in Python")

    live = sum(1 for _, _, k in COLUMNS if k == "formula")
    print(f"  {live} live formula columns, all arithmetic; no Poisson in the sheet")
    capped = int((d["matches"] >= xpts_calc.MATCHES - 1e-9).sum())
    print(f"  {capped} players pinned at the {xpts_calc.MATCHES}-match ceiling")
    return 0


if __name__ == "__main__":
    sys.exit(main())
